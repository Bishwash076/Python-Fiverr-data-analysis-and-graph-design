import openpyxl
from openpyxl.utils import get_column_letter
import uuid

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Methanol Mass Balance"

# Define the layout
# Main table: Streams 1-19
headers = ["Stream", "Total Flow (Nm³/h)", "CO2 (%)", "CO (%)", "H2 (%)", "Inert (%)", "H2O (%)", "MeOH (%)", "Mass Flow (kg/h)"]
for col, header in enumerate(headers, start=1):
    ws[f"{get_column_letter(col)}1"] = header

# Stream numbers
for row in range(2, 21):
    ws[f"A{row}"] = row - 1

# Molecular Weights (columns K and L)
mw_data = {
    "CO2": 44.01,
    "CO": 28.01,
    "H2": 2.02,
    "Inert": 39.94,
    "H2O": 18.015,
    "MeOH": 32.04
}
ws["K1"] = "Component"
ws["L1"] = "MW (g/mol)"
for i, (comp, mw) in enumerate(mw_data.items(), start=2):
    ws[f"K{i}"] = comp
    ws[f"L{i}"] = mw

# Feed Composition (columns R and S)
feed_comp = {"CO2": 0.33, "CO": 0.34, "H2": 0.31, "Inert": 0.02}
ws["R1"] = "Component"
ws["S1"] = "Fraction"
for i, (comp, frac) in enumerate(feed_comp.items(), start=2):
    ws[f"R{i}"] = comp
    ws[f"S{i}"] = frac

# Process Conditions (columns U and V)
process_conditions = {
    "feed_flow": 1759,
    "target_SN": 2.1,
    "shift_conversion": 0.98,
    "co2_removal_rate": 0.95,
    "inert_removal_rate": 0.99,
    "h2_retention": 0.93,
    "methanol_single_pass_conversion": 0.20,
    "methanol_overall_conversion": 0.95
}
ws["U1"] = "Parameter"
ws["V1"] = "Value"
for i, (param, value) in enumerate(process_conditions.items(), start=2):
    ws[f"U{i}"] = param
    ws[f"V{i}"] = value

# Split Ratio (approximated from original code)
ws["U9"] = "split_ratio"
ws["V9"] = 0.3057  # Pre-calculated value

# Helper function to generate mass flow formula
def mass_flow_formula(row):
    return (f"=B{row}*(C{row}/100*L2 + D{row}/100*L3 + E{row}/100*L4 + "
            f"F{row}/100*L5 + G{row}/100*L6 + H{row}/100*L7)")

# Populate formulas for each stream
for row in range(2, 21):
    stream = row - 1
    if stream == 1:  # Stream 1
        ws[f"B{row}"] = "=V2"
        ws[f"C{row}"] = "=S2*100"
        ws[f"D{row}"] = "=S3*100"
        ws[f"E{row}"] = "=S4*100"
        ws[f"F{row}"] = "=S5*100"
        ws[f"G{row}"] = "=0"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 2:  # Stream 2
        ws[f"B{row}"] = "=V2*V9"
        ws[f"C{row}"] = "=C2"
        ws[f"D{row}"] = "=D2"
        ws[f"E{row}"] = "=E2"
        ws[f"F{row}"] = "=F2"
        ws[f"G{row}"] = "=0"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 3:  # Stream 3
        ws[f"B{row}"] = "=V2*(1-V9)"
        ws[f"C{row}"] = "=C2"
        ws[f"D{row}"] = "=D2"
        ws[f"E{row}"] = "=E2"
        ws[f"F{row}"] = "=F2"
        ws[f"G{row}"] = "=0"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 4:  # Stream 4
        ws[f"B{row}"] = "=B4+B6"
        ws[f"C{row}"] = "=(C4*B4)/B5"
        ws[f"D{row}"] = "=(D4*B4)/B5"
        ws[f"E{row}"] = "=(E4*B4)/B5"
        ws[f"F{row}"] = "=(F4*B4)/B5"
        ws[f"G{row}"] = "=100*B6/B5"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 5:  # Stream 5
        ws[f"B{row}"] = "=B4*S3"
        ws[f"C{row}"] = "=0"
        ws[f"D{row}"] = "=0"
        ws[f"E{row}"] = "=0"
        ws[f"F{row}"] = "=0"
        ws[f"G{row}"] = "=100"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 6:  # Stream 6
        ws[f"B{row}"] = "=B5"
        ws[f"C{row}"] = "=C5 + 100*(D5/100*B5*V4)/B7"
        ws[f"D{row}"] = "=100*(D5/100*B5*(1-V4))/B7"
        ws[f"E{row}"] = "=E5 + 100*(D5/100*B5*V4)/B7"
        ws[f"F{row}"] = "=F5"
        ws[f"G{row}"] = "=G5 + 100*(D5/100*B5*V4)/B7"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 7:  # Stream 7
        ws[f"B{row}"] = "=B7*G7/100"
        ws[f"C{row}"] = "=0"
        ws[f"D{row}"] = "=0"
        ws[f"E{row}"] = "=0"
        ws[f"F{row}"] = "=0"
        ws[f"G{row}"] = "=100"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 8:  # Stream 8
        ws[f"B{row}"] = "=B7-B8"
        ws[f"C{row}"] = "=(C7/(100-G7))*100"
        ws[f"D{row}"] = "=(D7/(100-G7))*100"
        ws[f"E{row}"] = "=(E7/(100-G7))*100"
        ws[f"F{row}"] = "=(F7/(100-G7))*100"
        ws[f"G{row}"] = "=0"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 9:  # Stream 9
        ws[f"B{row}"] = "=(C9/100)*B9*V5"
        ws[f"C{row}"] = "=100"
        ws[f"D{row}"] = "=0"
        ws[f"E{row}"] = "=0"
        ws[f"F{row}"] = "=0"
        ws[f"G{row}"] = "=0"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 10:  # Stream 10
        ws[f"B{row}"] = "=SUMPRODUCT(B9*{C9,D9,E9,F9}/100*{1-V5,1,1,1})"
        ws[f"C{row}"] = "=100*(C9/100*B9*(1-V5))/B11"
        ws[f"D{row}"] = "=100*(D9/100*B9)/B11"
        ws[f"E{row}"] = "=100*(E9/100*B9)/B11"
        ws[f"F{row}"] = "=100*(F9/100*B9)/B11"
        ws[f"G{row}"] = "=0"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 11:  # Stream 11 (simplified, assuming same composition as Stream 10)
        ws[f"B{row}"] = "=E10/100*B10*(1-V6) + SUMPRODUCT(B10*{C10,D10,F10}/100*V7)"
        ws[f"C{row}"] = "=C10"
        ws[f"D{row}"] = "=D10"
        ws[f"E{row}"] = "=E10"
        ws[f"F{row}"] = "=F10"
        ws[f"G{row}"] = "=0"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 12:  # Stream 12
        ws[f"B{row}"] = "=SUMPRODUCT(B10*{C10,D10,E10,F10}/100*{1-V7,1-V7,V6,1-V7})"
        ws[f"C{row}"] = "=100*(C10/100*B10*(1-V7))/B13"
        ws[f"D{row}"] = "=100*(D10/100*B10*(1-V7))/B13"
        ws[f"E{row}"] = "=100*(E10/100*B10*V6)/B13"
        ws[f"F{row}"] = "=100*(F10/100*B10*(1-V7))/B13"
        ws[f"G{row}"] = "=0"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 13:  # Stream 13
        ws[f"B{row}"] = "=B3+B13"
        ws[f"C{row}"] = "=(C3*B3 + C13*B13)/B14"
        ws[f"D{row}"] = "=(D3*B3 + D13*B13)/B14"
        ws[f"E{row}"] = "=(E3*B3 + E13*B13)/B14"
        ws[f"F{row}"] = "=(F3*B3 + F13*B13)/B14"
        ws[f"G{row}"] = "=0"
        ws[f"H{row}"] = "=0"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 14:  # Stream 14
        for col in range(2, 10):
            ws[f"{get_column_letter(col)}{row}"] = f"={get_column_letter(col)}{row-1}"

    elif stream == 15:  # Stream 15
        ws[f"B{row}"] = "=B14"
        ws[f"C{row}"] = "=C14*(1-V7)"
        ws[f"D{row}"] = "=D14*(1-V7)"
        ws[f"E{row}"] = "=E14*(1-V7)"
        ws[f"F{row}"] = "=F14"
        ws[f"G{row}"] = "=0"
        ws[f"H{row}"] = "=100*V7"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 16:  # Stream 16
        ws[f"B{row}"] = "=B14*V8"
        ws[f"C{row}"] = "=0"
        ws[f"D{row}"] = "=0"
        ws[f"E{row}"] = "=0"
        ws[f"F{row}"] = "=0"
        ws[f"G{row}"] = "=0"
        ws[f"H{row}"] = "=100"
        ws[f"I{row}"] = mass_flow_formula(row)

    elif stream == 17:  # Stream 18
        for col in range(2, 10):
            ws[f"{get_column_letter(col)}{row}"] = f"={get_column_letter(col)}{row-2}"

    elif stream == 18:  # Stream 19
        ws[f"B{row}"] = "=B14*F14/100"
        ws[f"C{row}"] = "=C16"
        ws[f"D{row}"] = "=D16"
        ws[f"E{row}"] = "=E16"
        ws[f"F{row}"] = "=F16"
        ws[f"G{row}"] = "=G16"
        ws[f"H{row}"] = "=H16"
        ws[f"I{row}"] = mass_flow_formula(row)

# Add SN check for Stream 13
ws["J1"] = "SN (Stream 13)"
ws["J14"] = "=E14/(D14 + 3*C14)"

# Save the Excel file
excel_name = "Methanol_Mass_Balance_Formulas.xlsx"
wb.save(excel_name)

# Download in Google Colab or print confirmation
try:
    from google.colab import files
    files.download(excel_name)
    print(f"Excel file '{excel_name}' generated and downloading...")
except ImportError:
    print(f"✔ Excel file '{excel_name}' saved in the current directory. You can download it from there.")