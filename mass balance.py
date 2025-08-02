import openpyxl 
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Methanol Mass Balance"

# Define the layout
# Main table: Streams 1-19 with compound triplets
headers = [
    "Stream", "Total Flow (Nm³/h)",
    "CO2 (%)", "CO2 Flow (Nm³/h)", "CO2 Mass (kg/h)",
    "CO (%)", "CO Flow (Nm³/h)", "CO Mass (kg/h)",
    "H2 (%)", "H2 Flow (Nm³/h)", "H2 Mass (kg/h)",
    "Inert (%)", "Inert Flow (Nm³/h)", "Inert Mass (kg/h)",
    "H2O (%)", "H2O Flow (Nm³/h)", "H2O Mass (kg/h)",
    "MeOH (%)", "MeOH Flow (Nm³/h)", "MeOH Mass (kg/h)",
    "Total Mass Flow (kg/h)", "Sum of Mass Flows (kg/h)"
]
for col, header in enumerate(headers, start=1):
    ws[f"{get_column_letter(col)}1"] = header

# Stream numbers
for row in range(2, 21):
    ws[f"A{row}"] = row - 1

# Molecular Weights (columns W and X)
mw_data = {
    "CO2": 44.01,
    "CO": 28.01,
    "H2": 2.02,
    "Inert": 39.94,
    "H2O": 18.015,
    "MeOH": 32.04
}
ws["W1"] = "Component"
ws["X1"] = "MW (g/mol)"
for i, (comp, mw) in enumerate(mw_data.items(), start=2):
    ws[f"W{i}"] = comp
    ws[f"X{i}"] = mw

# Molar volume constant (22.414 L/mol at STP, converted to Nm³)
ws["W8"] = "Molar Volume (Nm³/mol)"
ws["X8"] = 0.022414

# Feed Composition (columns Z and AA)
feed_comp = {"CO2": 0.33, "CO": 0.34, "H2": 0.31, "Inert": 0.02}
ws["Z1"] = "Component"
ws["AA1"] = "Fraction"
for i, (comp, frac) in enumerate(feed_comp.items(), start=2):
    ws[f"Z{i}"] = comp
    ws[f"AA{i}"] = frac

# Process Conditions (columns AC and AD)
process_conditions = {
    "feed_flow": 1759,
    "target_SN": 2.1,
    "shift_conversion": 0.98,
    "co2_removal_rate": 0.95,
    "inert_removal_rate": 0.99,
    "h2_retention": 0.93,
    "methanol_single_pass_conversion": 0.20,
    "methanol_overall_conversion": 0.95
}
ws["AC1"] = "Parameter"
ws["AD1"] = "Value"
for i, (param, value) in enumerate(process_conditions.items(), start=2):
    ws[f"AC{i}"] = param
    ws[f"AD{i}"] = value

# Split Ratio (approximated from original code)
ws["AC9"] = "split_ratio"
ws["AD9"] = 0.3057  # Pre-calculated value

# Helper functions for formulas
def component_flow_formula(row, comp_col):
    return f"=B{row}*{get_column_letter(comp_col)}{row}/100"

def component_mass_formula(row, flow_col, mw_row):
    return f"={get_column_letter(flow_col)}{row}*X{mw_row}/X8"

def total_mass_flow_formula(row):
    return f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

# Populate formulas for each stream
for row in range(2, 21):
    stream = row - 1
    if stream == 1:  # Stream 1
        ws[f"B{row}"] = "=AD2"
        ws[f"C{row}"] = "=AA2*100"  # CO2 %
        ws[f"D{row}"] = component_flow_formula(row, 3)  # CO2 Flow
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)  # CO2 Mass
        ws[f"F{row}"] = "=AA3*100"  # CO %
        ws[f"G{row}"] = component_flow_formula(row, 6)  # CO Flow
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)  # CO Mass
        ws[f"I{row}"] = "=AA4*100"  # H2 %
        ws[f"J{row}"] = component_flow_formula(row, 9)  # H2 Flow
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)  # H2 Mass
        ws[f"L{row}"] = "=AA5*100"  # Inert %
        ws[f"M{row}"] = component_flow_formula(row, 12)  # Inert Flow
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)  # Inert Mass
        ws[f"O{row}"] = "=0"  # H2O %
        ws[f"P{row}"] = component_flow_formula(row, 15)  # H2O Flow
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)  # H2O Mass
        ws[f"R{row}"] = "=0"  # MeOH %
        ws[f"S{row}"] = component_flow_formula(row, 18)  # MeOH Flow
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)  # MeOH Mass
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 2:  # Stream 2
        ws[f"B{row}"] = "=AD2*AD9"
        ws[f"C{row}"] = "=C2"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=F2"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=I2"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=L2"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=O2"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=R2"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 3:  # Stream 3
        ws[f"B{row}"] = "=AD2*(1-AD9)"
        ws[f"C{row}"] = "=C2"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=F2"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=I2"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=L2"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=O2"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=R2"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 4:  # Stream 4
        ws[f"B{row}"] = "=B4+B6"
        ws[f"C{row}"] = "=(C4*B4)/B5"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=(F4*B4)/B5"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=(I4*B4)/B5"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=(L4*B4)/B5"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=100*B6/B5"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=0"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 5:  # Stream 5
        ws[f"B{row}"] = "=B4*AA3"
        ws[f"C{row}"] = "=0"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=0"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=0"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=0"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=100"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=0"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 6:  # Stream 6
        ws[f"B{row}"] = "=B5"
        ws[f"C{row}"] = "=C5 + 100*(F5/100*B5*AD4)/B7"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=100*(F5/100*B5*(1-AD4))/B7"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=I5 + 100*(F5/100*B5*AD4)/B7"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=L5"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=O5 + 100*(F5/100*B5*AD4)/B7"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=0"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 7:  # Stream 7
        ws[f"B{row}"] = "=B7*O7/100"
        ws[f"C{row}"] = "=0"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=0"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=0"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=0"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=100"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=0"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 8:  # Stream 8
        ws[f"B{row}"] = "=B7-B8"
        ws[f"C{row}"] = "=(C7/(100-O7))*100"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=(F7/(100-O7))*100"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=(I7/(100-O7))*100"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=(L7/(100-O7))*100"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=0"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=0"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 9:  # Stream 9
        ws[f"B{row}"] = "=(C9/100)*B9*AD5"
        ws[f"C{row}"] = "=100"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=0"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=0"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=0"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=0"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=0"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 10:  # Stream 10
        ws[f"B{row}"] = "=SUMPRODUCT(B9*{C9,F9,I9,L9}/100*{1-AD5,1,1,1})"
        ws[f"C{row}"] = "=100*(C9/100*B9*(1-AD5))/B11"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=100*(F9/100*B9)/B11"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=100*(I9/100*B9)/B11"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=100*(L9/100*B9)/B11"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=0"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=0"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 11:  # Stream 11
        ws[f"B{row}"] = "=I10/100*B10*(1-AD6) + SUMPRODUCT(B10*{C10,F10,L10}/100*AD7)"
        ws[f"C{row}"] = "=C10"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=F10"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=I10"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=L10"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=0"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=0"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 12:  # Stream 12
        ws[f"B{row}"] = "=SUMPRODUCT(B10*{C10,F10,I10,L10}/100*{1-AD7,1-AD7,AD6,1-AD7})"
        ws[f"C{row}"] = "=100*(C10/100*B10*(1-AD7))/B13"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=100*(F10/100*B10*(1-AD7))/B13"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=100*(I10/100*B10*AD6)/B13"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=100*(L10/100*B10*(1-AD7))/B13"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=0"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=0"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 13:  # Stream 13
        ws[f"B{row}"] = "=B3+B13"
        ws[f"C{row}"] = "=(C3*B3 + C13*B13)/B14"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=(F3*B3 + F13*B13)/B14"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=(I3*B3 + I13*B13)/B14"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=(L3*B3 + L13*B13)/B14"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=0"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=0"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 14:  # Stream 14
        for col in range(2, 19):
            ws[f"{get_column_letter(col)}{row}"] = f"={get_column_letter(col)}{row-1}"
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 15:  # Stream 15
        ws[f"B{row}"] = "=B14"
        ws[f"C{row}"] = "=C14*(1-AD7)"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=F14*(1-AD7)"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=I14*(1-AD7)"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=L14"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=0"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=100*AD7"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 16:  # Stream 16
        ws[f"B{row}"] = "=B14*AD8"
        ws[f"C{row}"] = "=0"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=0"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=0"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=0"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=0"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=100"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 17:  # Stream 17
        ws[f"B{row}"] = "=0"  # Unused stream in process
        ws[f"C{row}"] = "=0"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=0"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=0"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=0"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=0"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=0"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 18:  # Stream 18
        for col in range(2, 19):
            ws[f"{get_column_letter(col)}{row}"] = f"={get_column_letter(col)}{row-2}"
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

    elif stream == 19:  # Stream 19
        ws[f"B{row}"] = "=B14*L14/100"
        ws[f"C{row}"] = "=C18"
        ws[f"D{row}"] = component_flow_formula(row, 3)
        ws[f"E{row}"] = component_mass_formula(row, 4, 2)
        ws[f"F{row}"] = "=F18"
        ws[f"G{row}"] = component_flow_formula(row, 6)
        ws[f"H{row}"] = component_mass_formula(row, 7, 3)
        ws[f"I{row}"] = "=I18"
        ws[f"J{row}"] = component_flow_formula(row, 9)
        ws[f"K{row}"] = component_mass_formula(row, 10, 4)
        ws[f"L{row}"] = "=L18"
        ws[f"M{row}"] = component_flow_formula(row, 12)
        ws[f"N{row}"] = component_mass_formula(row, 13, 5)
        ws[f"O{row}"] = "=O18"
        ws[f"P{row}"] = component_flow_formula(row, 15)
        ws[f"Q{row}"] = component_mass_formula(row, 16, 6)
        ws[f"R{row}"] = "=R18"
        ws[f"S{row}"] = component_flow_formula(row, 18)
        ws[f"T{row}"] = component_mass_formula(row, 19, 7)
        ws[f"U{row}"] = total_mass_flow_formula(row)
        ws[f"V{row}"] = f"=SUM(E{row},H{row},K{row},N{row},Q{row},T{row})"

# Add SN check for Stream 13
ws["W13"] = "SN (Stream 13)"
ws["W14"] = "=I14/(F14 + 3*C14)"

# Add inert balance check
ws["W15"] = "Inert Check (Stream 1 = Stream 19)"
ws["W16"] = "=B2*L2/100"
ws["W17"] = "=B20*L20/100"

# Add notes about iterative calculations and unused streams
ws["W19"] = "Note: Enable iterative calculations in Excel (File > Options > Formulas) to resolve recycle loop."
ws["W20"] = "Note: Stream 17 is unused in the process, hence all values are set to 0."

# Save the Excel file
excel_name = "Methanol_Mass_Balance_Formulas.xlsx"
wb.save(excel_name)

# Download in Google Colab or print confirmation
try:
    from google.colab import files
    files.download(excel_name)
    print(f"Excel file '{excel_name}' generated and downloading...")
except ImportError:
    print(f"✔ Excel file '{excel_name}' saved in the current directory. You can download it from there.")
